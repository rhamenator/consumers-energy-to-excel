from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import process_consumers_itemized_statement as pcs


def _rgb(cell) -> str:
    rgb = cell.fill.fgColor.rgb
    return rgb or ""


def test_apply_excel_formatting_greenbar_header_and_freeze(tmp_path: Path):
    df = pcs.pd.DataFrame(
        [
            {
                "Due Date": "01/12/2023",
                "Total Bill": "$10.00",
                "Payment Received": "$1.00-",
            },
            {
                "Due Date": "01/13/2023",
                "Total Bill": "$20.00",
                "Payment Received": "$2.00-",
            },
        ]
    )

    normalized, hints = pcs.normalize_dataframe(df)
    output = tmp_path / "out.xlsx"
    normalized.to_excel(output, index=False)
    pcs.apply_excel_formatting(output, hints)

    wb = load_workbook(output)
    ws = wb.active

    # Header row is gray
    assert _rgb(ws["A1"]).endswith("D9D9D9")

    # Greenbar alternating: row 2 (first data) white-ish (no fill), row 3 pale green
    assert _rgb(ws["A3"]).endswith("E8F3E8")

    # Freeze panes at A2
    assert ws.freeze_panes == "A2"

    # Accounting format applied to Total Bill
    total_bill_cell = ws["B2"]
    assert "$" in (total_bill_cell.number_format or "")

    # Column widths set (at least the minimum)
    assert ws.column_dimensions["A"].width >= 12
