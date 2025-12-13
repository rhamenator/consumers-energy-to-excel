import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
import tkinter as tk
from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox

try:
    import pdfplumber
except ImportError:  # pragma: no cover - optional dependency warning shown later
    pdfplumber = None


COLUMN_NAMES = [
    "Current Bill Month-Year",
    "Due Date",
    "Meter Number",
    "Begin Date",
    "End Date",
    "Days Billed",
    "Begin Read",
    "End Read",
    "Read Type",
    "Energy Used",
    "Use Per Day",
    "Power Factor",
    "Billed KW",
    "Max KW",
    "Balance After Current Charges",
    "Winter Protection Payment",
    "Monthly Charge",
    "Tax",
    "Total Bill",
    "Payment Received",
    "Payment Date",
    "Late Charge",
    "Adjustment Date",
    "Last Month Balance",
    "Balance Before Current Charges",
]

# Keywords used to identify electric billing rows for electric customers
ELECTRIC_KEYWORDS = [
    "electric",
    "electricity",
    "peak",
    "on-peak",
    "off-peak",
    "weekend",
    "weekday",
]


def main() -> None:
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Itemized Statement (PDF or XLSX)",
        filetypes=(
            ("PDF and XLSX files", "*.pdf *.xlsx"),
            ("PDF files", "*.pdf"),
            ("XLSX files", "*.xlsx"),
            ("All files", "*.*"),
        ),
    )

    if not file_path:
        return

    try:
        itemized_statement_df = load_itemized_statement(Path(file_path))
    except Exception as exc:  # pragma: no cover - surfaced to the user
        messagebox.showerror("Error", f"Unable to load statement: {exc}")
        return

    mapped_df = map_monthly_data(itemized_statement_df)
    normalized_df, format_hints = normalize_dataframe(mapped_df)
    output_path = Path("output.xlsx")
    csv_path = Path("output.csv")
    normalized_df.to_excel(output_path, index=False)
    normalized_df.to_csv(csv_path, index=False)
    apply_excel_formatting(output_path, format_hints)
    messagebox.showinfo(
        "Success",
        "Mapped data saved to output.xlsx and output.csv",
    )


def load_itemized_statement(file_path: Path) -> pd.DataFrame:
    extension = file_path.suffix.lower()
    if extension == ".pdf":
        if pdfplumber is None:
            raise ImportError("pdfplumber is required to process PDF statements.")
        raw_df = read_pdf_tables(file_path)
    elif extension in {".xlsx", ".xls"}:
        raw_df = pd.read_excel(file_path, header=None)
    else:
        raise ValueError("Please select a PDF or XLSX file.")

    cleaned_df = clean_itemized_statement(raw_df)
    if "Description" not in cleaned_df.columns:
        raise ValueError("Unable to find a Description column in the statement.")
    cleaned_df["Description"] = (
        cleaned_df["Description"].astype(str).str.replace("\n", " ")
    )
    if "Use Per Day" not in cleaned_df.columns and "Use Per" in cleaned_df.columns:
        cleaned_df = cleaned_df.rename(columns={"Use Per": "Use Per Day"})
    elif "Use Per Day" not in cleaned_df.columns:
        cleaned_df["Use Per Day"] = None

    return cleaned_df


def read_pdf_tables(pdf_path: Path) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    current_header: list[str] | None = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                df = pd.DataFrame(table)
                df = _standardize_cells(df)
                if df.empty:
                    continue
                if _row_looks_like_header(df.iloc[0]):
                    current_header = df.iloc[0].tolist()
                    data = df.iloc[1:].copy()
                else:
                    data = df.copy()
                if current_header is None or data.empty:
                    continue
                data.columns = current_header
                frames.append(data)

    if not frames:
        raise ValueError("No recognizable tables were found in the PDF.")

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.replace({"" : pd.NA})
    combined = combined.dropna(how="all")
    return combined


def _standardize_cells(df: pd.DataFrame) -> pd.DataFrame:
    df = df.replace({"\u00a0": " ", "\n": " "}, regex=True)
    df = df.map(lambda value: value.strip() if isinstance(value, str) else value)
    return df


def _table_contains_key(df: pd.DataFrame, keyword: str) -> bool:
    keyword_lower = keyword.lower()
    for _, row in df.iterrows():
        for value in row:
            if isinstance(value, str) and keyword_lower in value.lower():
                return True
    return False


def _row_looks_like_header(row: pd.Series) -> bool:
    values = [str(value).strip().lower() for value in row.tolist() if isinstance(value, str)]
    return "description" in values and "monthly charge" in values


def clean_itemized_statement(raw_df: pd.DataFrame) -> pd.DataFrame:
    sanitized = raw_df.copy()
    sanitized = sanitized.replace({"\u00a0": " "})
    sanitized = sanitized.dropna(how="all")
    if "Description" in sanitized.columns:
        data = sanitized.copy()
    else:
        header_row_idx = find_header_row_index(sanitized)
        header_row = (
            sanitized.iloc[header_row_idx]
            .astype(str)
            .str.replace("\n", " ")
            .str.strip()
            .tolist()
        )
        data = sanitized.iloc[header_row_idx + 1 :].copy()
        data.columns = header_row
    data = data.dropna(how="all")
    data = data.reset_index(drop=True)
    if "Description" in data.columns:
        data["Description"] = data["Description"].astype(str).str.strip()
        data = data[data["Description"].str.len() > 0]
    return data


def find_header_row_index(df: pd.DataFrame) -> int:
    for idx, row in df.iterrows():
        values = [str(value).strip().lower() for value in row.tolist() if value is not None]
        if "description" in values and "monthly charge" in values:
            return idx
    return 0


def map_monthly_data(itemized_statement_df: pd.DataFrame) -> pd.DataFrame:
    mapped_rows: list[dict[str, object]] = []
    for month_df in split_into_months(itemized_statement_df):
        if month_df.empty:
            continue
        mapped_row = map_single_month(month_df)
        if mapped_row:
            mapped_rows.append(mapped_row)
    if not mapped_rows:
        return pd.DataFrame(columns=COLUMN_NAMES)
    return pd.DataFrame(mapped_rows, columns=COLUMN_NAMES)


def split_into_months(df: pd.DataFrame) -> list[pd.DataFrame]:
    months: list[pd.DataFrame] = []
    buffer: list[dict[str, object]] = []
    for _, row in df.iterrows():
        desc = str(row.get("Description", "")).strip()
        if not desc or desc.lower() == "nan":
            continue
        buffer.append(row.to_dict())
        normalized = desc.lower().replace(" ", "")
        if "totalamountdue" in normalized:
            months.append(pd.DataFrame(buffer, columns=df.columns).reset_index(drop=True))
            buffer = []
    if buffer:
        months.append(pd.DataFrame(buffer, columns=df.columns).reset_index(drop=True))
    return months


def map_single_month(month_data: pd.DataFrame) -> dict | None:
    mapped_data: dict[str, object] = {}
    try:
        begin_date_row = _find_energy_row(month_data)
        if begin_date_row is None:
            raise ValueError("Missing Gas or Electric row for month")
        mapped_data["Current Bill Month-Year"] = pd.to_datetime(
            begin_date_row["Begin Date"]
        ).strftime("%Y-%m-%d")
        due_row = _first_row(month_data, "due on or before")
        if due_row is None:
            raise ValueError("Missing due date row for month")
        mapped_data["Due Date"] = pd.to_datetime(
            str(due_row.get("Description", "")).split()[-1],
            format="%m/%d/%Y",
        ).strftime("%Y-%m-%d")
        mapped_data["Meter Number"] = begin_date_row.get("Meter Number")
        mapped_data["Begin Date"] = begin_date_row.get("Begin Date")
        mapped_data["End Date"] = begin_date_row.get("End Date")
        mapped_data["Days Billed"] = begin_date_row.get("Days Billed")
        mapped_data["Begin Read"] = begin_date_row.get("Begin Read")
        mapped_data["End Read"] = begin_date_row.get("End Read")
        mapped_data["Read Type"] = begin_date_row.get("Read Type")
        mapped_data["Energy Used"] = begin_date_row.get("Energy Used")
        mapped_data["Use Per Day"] = begin_date_row.get("Use Per Day")
        mapped_data["Power Factor"] = begin_date_row.get("Power Factor")
        mapped_data["Billed KW"] = begin_date_row.get("Billed KW")
        mapped_data["Max KW"] = begin_date_row.get("Max KW")
        mapped_data["Balance After Current Charges"] = _first_value(
            month_data,
            "BALANCE AFTER CURRENT CHARGES",
            "Monthly Charge",
        )
        mapped_data["Winter Protection Payment"] = _first_value(
            month_data,
            "WINTER PROTECTION PLAN PAYMENT",
            "Monthly Charge",
        )
        mapped_data["Monthly Charge"] = begin_date_row.get("Monthly Charge")
        mapped_data["Tax"] = _exact_value(month_data, "Tax", "Monthly Charge")
        mapped_data["Total Bill"] = _first_value(
            month_data,
            "total amount due",
            "Account Balance",
        )
        mapped_data["Payment Received"] = _first_value(
            month_data,
            "Payment Received",
            "Transaction Amt",
        )
        mapped_data["Payment Date"] = _first_value(
            month_data,
            "Payment Received",
            "Date",
        )
        mapped_data["Late Charge"] = _first_value(
            month_data,
            "LATE PAYMENT CHARGE",
            "Transaction Amt",
        )
        canceled_row = _first_row(month_data, "Canceled Late Payment Charge")
        if canceled_row is not None:
            mapped_data["Late Charge"] = canceled_row.get("Transaction Amt")
            mapped_data["Adjustment Date"] = canceled_row.get("Date")
        mapped_data["Last Month Balance"] = _first_value(
            month_data,
            "Last Month's Consumers Energy Account Balance",
            "Account Balance",
        )
        mapped_data["Balance Before Current Charges"] = _first_value(
            month_data,
            "Account Balance Before Current Charges",
            "Account Balance",
        )
        return mapped_data
    except Exception as exc:  # pragma: no cover - log and skip malformed month
        print(f"Skipping a month due to error: {exc}", file=sys.stderr)
        return None


def _find_energy_row(df: pd.DataFrame) -> pd.Series | None:
    """Find the main energy row (gas or electric) in the month data.

    Tries to find a row with gas first, then searches for electricity-related rows.
    Electric rows may contain words like: electric, electricity, peak, on-peak,
    off-peak, weekend, weekday, etc.
    """
    # First try to find a gas row (for gas customers)
    gas_row = _first_row(df, "gas")
    if gas_row is not None:
        return gas_row
    
    # If no gas row, try to find electric-related rows
    for keyword in ELECTRIC_KEYWORDS:
        electric_row = _first_row(df, keyword)
        if electric_row is not None:
            return electric_row
    
    return None


def _first_row(df: pd.DataFrame, pattern: str):
    if "Description" not in df.columns:
        return None
    pattern_clean = pattern.lower().replace(" ", "")
    descriptions = (
        df["Description"]
        .astype(str)
        .str.lower()
        .str.replace(r"\s+", "", regex=True)
    )
    matching_mask = descriptions.str.contains(pattern_clean, na=False)
    matching = df[matching_mask]
    return matching.iloc[0] if not matching.empty else None


def _first_value(df: pd.DataFrame, pattern: str, column: str):
    row = _first_row(df, pattern)
    if row is None:
        return None
    return row.get(column)


def _exact_value(df: pd.DataFrame, exact_description: str, column: str):
    if "Description" not in df.columns:
        return None
    matching = df[df["Description"] == exact_description]
    if matching.empty:
        return None
    row = matching.iloc[0]
    return row.get(column)


DATE_PATTERN = re.compile(r"^\d{1,4}[/-]\d{1,2}[/-]\d{1,4}$")
NUMBER_PATTERN = re.compile(r"^\d+(\.\d+)?$")


def normalize_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[tuple[int, int], dict[str, Any]]]:
    if df.empty:
        return df.copy(), {}

    normalized_rows: list[list[Any]] = []
    format_hints: dict[tuple[int, int], dict[str, Any]] = {}

    for row_idx, (_, row) in enumerate(df.iterrows()):
        normalized_row: list[Any] = []
        for col_idx, value in enumerate(row):
            normalized_value, fmt = normalize_cell(value)
            normalized_row.append(normalized_value)
            if fmt is not None:
                format_hints[(row_idx, col_idx)] = fmt
        normalized_rows.append(normalized_row)

    normalized_df = pd.DataFrame(normalized_rows, columns=df.columns)
    return normalized_df, format_hints


def normalize_cell(value: Any) -> tuple[Any, dict[str, Any] | None]:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None, None

    if isinstance(value, (datetime, date)):
        return value, {"type": "date"}

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        decimals = _infer_decimal_places_from_number(value)
        value_cast = float(value) if isinstance(value, float) else int(value)
        fmt_type = "number"
        return value_cast, {"type": fmt_type, "decimals": decimals}

    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"none", "nan"}:
            return None, None

        if DATE_PATTERN.fullmatch(text):
            try:
                parsed_date = date_parser.parse(text)
                return parsed_date.date(), {"type": "date"}
            except (ValueError, OverflowError):
                pass

        numeric_value, fmt = _parse_numeric(text)
        if numeric_value is not None:
            return numeric_value, fmt

        return text, None

    return value, None


def _infer_decimal_places_from_number(value: float | int) -> int:
    if isinstance(value, int):
        return 0
    decimal_value = Decimal(str(value)).normalize()
    return max(-decimal_value.as_tuple().exponent, 0)


def _parse_numeric(text: str) -> tuple[Any | None, dict[str, Any] | None]:
    stripped = text.replace(",", "").replace(" ", "")
    accounting = stripped.startswith("$")
    if accounting:
        stripped = stripped[1:]

    negative = stripped.endswith("-")
    if negative:
        stripped = stripped[:-1]

    if stripped.startswith("-"):
        negative = True
        stripped = stripped[1:]

    if not stripped:
        return None, None

    if NUMBER_PATTERN.fullmatch(stripped):
        decimals = len(stripped.split(".")[1]) if "." in stripped else 0
        number = float(stripped) if "." in stripped else int(stripped)
        if negative:
            number = -number
        fmt_type = "accounting" if accounting else "number"
        return number, {"type": fmt_type, "decimals": decimals}

    return None, None


def apply_excel_formatting(output_path: Path, format_hints: dict[tuple[int, int], dict[str, Any]]) -> None:
    if not output_path.exists():
        return

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    green_fill = PatternFill(fill_type="solid", fgColor="E8F3E8")
    for row_idx in range(2, worksheet.max_row + 1):
        data_idx = row_idx - 2
        if data_idx % 2 == 1:
            for cell in worksheet[row_idx]:
                cell.fill = green_fill

    for (row_idx, col_idx), fmt in format_hints.items():
        excel_row = row_idx + 2  # offset for header
        excel_col = col_idx + 1
        cell = worksheet.cell(row=excel_row, column=excel_col)
        fmt_type = fmt.get("type")
        if fmt_type == "date":
            cell.number_format = "mm/dd/yyyy"
        elif fmt_type == "accounting":
            decimals = fmt.get("decimals", 2)
            cell.number_format = build_accounting_format(decimals)
        elif fmt_type == "number":
            decimals = fmt.get("decimals", 0)
            cell.number_format = build_number_format(decimals)

    worksheet.freeze_panes = worksheet["A2"]
    _autofit_columns(worksheet)

    workbook.save(output_path)


def build_number_format(decimals: int) -> str:
    decimal_section = "." + ("0" * decimals) if decimals > 0 else ""
    return f"#,##0{decimal_section}"


def build_accounting_format(decimals: int) -> str:
    decimals = max(decimals, 2)
    decimal_section = "." + ("0" * decimals)
    return (
        f"_($* #,##0{decimal_section}_);"
        f"_($* (#,##0{decimal_section});"
        f"_($* \"-\"??_);"
        "_(@_)"
    )


def _autofit_columns(worksheet) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in worksheet[column_letter]:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        adjusted_width = min(max(max_length + 2, 12), 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
if __name__ == "__main__":
    main()